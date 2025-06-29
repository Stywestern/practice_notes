import os

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill

from _config.settings import load_config
from _utils.functions import open_file

    
def compare() -> None:
    config = load_config()

    file1_path = config["service_two"]["file1"]["file_path"]
    sheet1 = config["service_two"]["file1"]["sheet_name"]

    file2_path = config["service_two"]["file2"]["file_path"]
    sheet2 = config["service_two"]["file2"]["sheet_name"]

    unique_key = config["service_two"]["unique_id_column"]
    compare_columns = config["service_two"]["compare_columns"]
    
    if unique_key in compare_columns:
        compare_columns.remove(unique_key)

    # Load the sheets into DataFrames
    df1 = pd.read_excel(file1_path, sheet_name=sheet1, engine='openpyxl')
    df2 = pd.read_excel(file2_path, sheet_name=sheet2, engine='openpyxl')
    
    ##### Align the dataframes
    # Step 1: Keep only intersecting columns and rows with common keys
    ### Cols
    common_columns = list(set(df1.columns).intersection(set(df2.columns)))
    if unique_key not in common_columns:
        print(f"❌ Unique key '{unique_key}' not found in both files.")
        return
    
    df1 = df1[common_columns].copy()
    df2 = df2[common_columns].copy()
    
    compare_columns = [col for col in compare_columns if col in common_columns]
    stable_columns = [col for col in common_columns if col not in compare_columns]

    # Defensive: ensure unique key is at the front
    if unique_key in stable_columns:
        stable_columns.remove(unique_key)
    stable_columns.insert(0, unique_key)

    # New column order: stable first, then compare columns
    reordered_columns = stable_columns + compare_columns
    df1 = df1[reordered_columns].copy()
    df2 = df2[reordered_columns].copy()
    
    ### Rows
    df1_keys = set(df1[unique_key].dropna())
    df2_keys = set(df2[unique_key].dropna())

    common_keys = df1_keys.intersection(df2_keys)

    df1_aligned = df1[df1[unique_key].isin(common_keys)].copy()
    df2_aligned = df2[df2[unique_key].isin(common_keys)].copy()
    
    print(f"\nAligned DataFrames: {len(df1_aligned)} rows are in both files.")

    # Step 2: Sort both DataFrames 
    # Columns we can use to sort — everything except comparison columns
    sortable_columns = [col for col in df1.columns if col not in compare_columns]

    # Defensive: ensure unique_key is first and present
    if unique_key in sortable_columns:
        sortable_columns.remove(unique_key)
    sortable_columns.insert(0, unique_key)

    # Step 2: Sort both DataFrames using safe columns
    df1_aligned = df1_aligned.sort_values(by=sortable_columns).reset_index(drop=True)
    df2_aligned = df2_aligned.sort_values(by=sortable_columns).reset_index(drop=True)

    ## Step 3: Compare the DataFrames
    result_df = df1_aligned.copy()

    # Collect (row_idx, column_name) for changed cells
    mismatches = []

    for idx in range(len(df1_aligned)):
        for col in compare_columns:
            val1 = df1_aligned.iloc[idx][col]
            val2 = df2_aligned.iloc[idx][col]

            # Compare only non-NaN values
            if pd.isna(val1) and pd.isna(val2):
                continue  # treat both NaNs as equal

            if val1 != val2:
                # Save the change in the result DataFrame as "old_value -> new_value"
                result_df.at[idx, col] = f"{val1} → {val2}"

                # Track mismatched cells for highlighting
                mismatches.append((idx, col))
    print(len(mismatches), "mismatches found.")
    
    # Step 4: Highlight mismatches in the result DataFrame        
    wb = Workbook()

    # 1. **Info Sheet**
    info_sheet = wb.active
    info_sheet.title = "Info"

    # Add file paths and comparison explanation
    info_sheet["A1"] = "File 1 Path:"
    info_sheet["B1"] = file1_path
    info_sheet["A2"] = "File 2 Path:"
    info_sheet["B2"] = file2_path
    info_sheet["A3"] = "Comparison Explanation:"
    info_sheet["B3"] = "'old_value → new_value' means the value from File 1 compared with the value from File 2"

    # 2. **Full Result Sheet**
    wb.create_sheet("Full Result")
    full_result_sheet = wb["Full Result"]

    # Write column headers first
    for col_idx, col_name in enumerate(result_df.columns, 1):
        full_result_sheet.cell(row=1, column=col_idx, value=col_name)

    # Write result_df into the "Full Result" sheet (starting from row 2)
    for r_idx, row in result_df.iterrows():
        for c_idx, value in enumerate(row, 1):
            full_result_sheet.cell(row=r_idx + 2, column=c_idx, value=value)

    # Apply yellow highlight to the full result sheet for mismatches
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    highlighted_cells = []  # To track the rows with highlighted cells

    # Highlight cells that have mismatches
    for row_idx, col_name in mismatches:
        col_idx = list(result_df.columns).index(col_name) + 1  # Excel is 1-based
        cell = full_result_sheet.cell(row=row_idx + 2, column=col_idx)  # +2 for header + 0-based idx
        cell.fill = yellow_fill
        highlighted_cells.append((row_idx + 2, col_idx))  # Store the row and column index

    # 3. **Filtered Result Sheet** (Extract rows with highlighted cells, removing empty rows)
    wb.create_sheet("Filtered Result")
    filtered_result_sheet = wb["Filtered Result"]

    # Write column headers first
    for col_idx, col_name in enumerate(result_df.columns, 1):
        filtered_result_sheet.cell(row=1, column=col_idx, value=col_name)

    # Copy rows with highlighted cells into the filtered sheet
    filtered_row_set = set([r for r, _ in highlighted_cells])

    for r_idx, row in result_df.iterrows():
        if r_idx + 2 in filtered_row_set:  # Check if the row is highlighted
            for c_idx, value in enumerate(row, 1):
                cell = filtered_result_sheet.cell(row=r_idx + 2, column=c_idx, value=value)
                # Highlight only the changed cells
                if (r_idx + 2, c_idx) in highlighted_cells:  # If the cell was highlighted
                    cell.fill = yellow_fill

    # Remove empty rows from filtered result sheet
    rows_to_delete = []
    max_row = filtered_result_sheet.max_row
    max_column = filtered_result_sheet.max_column

    for row_idx in range(2, max_row + 1):  # Start from row 2 (after header)
        is_empty = True
        for col_idx in range(1, max_column + 1):
            if filtered_result_sheet.cell(row=row_idx, column=col_idx).value is not None:
                is_empty = False
                break
        if is_empty:
            rows_to_delete.append(row_idx)

    for row_idx in reversed(rows_to_delete):  # Delete rows in reverse order to avoid index issues
        filtered_result_sheet.delete_rows(row_idx)

    # Save the workbook to a file
    file1_name = os.path.basename(file1_path).split('.')[0]  # Remove extension
    file2_name = os.path.basename(file2_path).split('.')[0]  # Remove extension

    # Create a more descriptive output file name
    output_path = f"outputs/compare_changes/comparison_({file1_name}_{file2_name}).xlsx"
    wb.save(output_path)

    print(f"\n✅ Comparison complete. Output saved to: {output_path}")

    # Open the output file
    open_file(output_path)
        